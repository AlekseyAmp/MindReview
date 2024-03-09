import typing
from dataclasses import dataclass
from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.application.review import interfaces


@dataclass
class ExcelManager(interfaces.IExcelManager):
    """
    Класс для работы с Excel документами.
    """

    def load_data(self, file: typing.BinaryIO) -> Worksheet:
        """
        Загружает и возвращает активный лист из Excel файла.

        :param file: Файл Excel для загрузки.

        :return: Активный лист Excel.
        """

        wb = load_workbook(file)
        ws = wb.active
        return ws

    def create_analyze_report(
        self,
        short_analyze: list,
        full_analyze: list
    ) -> str:
        """
        Создает Excel файл с данными анализа.

        :param short_analyze: Короткий анализ.
        :param full_analyze: Полный анализ.

        :return: Путь к созданному файлу Excel.
        """

        # Создаем новую книгу Excel и активный лист
        wb = Workbook()
        ws = wb.active

        # Создаем заголовки для короткого анализа
        short_headers = [
            "Дата проведения анализа",
            "Источник загрузки",
            "Ссылка на источник",
            "Количество отзывов",
            "Наиболее популярное настроение",
            "Топ 5 ключевых слов",
            "Топ 5 городов"
        ]
        # Заполняем заголовки для короткого анализа
        for col, header in enumerate(short_headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Создаем заголовки для полного анализа
        full_headers = [
            "Номер",
            "Отзыв",
            "Сентимент",
            "Ключевые слова",
            "Упоминания городов",
            "Упоминания возрастов"
        ]
        # Заполняем заголовки для полного анализа
        for col, header in enumerate(full_headers, start=1):
            ws.cell(row=len(short_analyze) + 3, column=col, value=header)

        border_styles = Border(
            left=Side(style="thin", color="C0C0C0"),
            right=Side(style="thin", color="C0C0C0"),
            top=Side(style="thin", color="C0C0C0"),
            bottom=Side(style="thin", color="C0C0C0")
        )

        # Применяем стили ко всем заголовкам
        header_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid"
        )
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = border_styles

        for row in (1, len(short_analyze) + 3):
            for col in range(1, len(short_headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border

        # Подготовка стилей для строк таблицы
        row_styles = [
            PatternFill(
                start_color="E2EFDA",
                end_color="E2EFDA",
                fill_type="solid"
            ),
            PatternFill(
                start_color="D9D9D9",
                end_color="D9D9D9",
                fill_type="solid"
            )
        ]
        cell_border = border_styles

        # Заполняем таблицу с данными короткого анализа
        for index, row_data in enumerate(short_analyze, start=2):
            for col_index, value in enumerate(row_data, start=1):
                ws.cell(row=index, column=col_index, value=value)
                ws.cell(
                    row=index,
                    column=col_index
                ).fill = row_styles[index % 2]
                ws.cell(row=index, column=col_index).border = cell_border

        # Заполняем таблицу с данными полного анализа
        for index, row_data in enumerate(
            full_analyze, start=len(short_analyze) + 4
        ):
            for col_index, value in enumerate(row_data, start=1):
                ws.cell(row=index, column=col_index, value=value)
                ws.cell(
                    row=index,
                    column=col_index
                ).fill = row_styles[index % 2]
                ws.cell(row=index, column=col_index).border = cell_border

        # Сохраняем файл во временное местоположение
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmpfile:
            wb.save(tmpfile.name)
            return tmpfile.name
